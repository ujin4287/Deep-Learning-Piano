import random
import cv2

import time

mark_ls_w = [24, 42, 71, 102, 119, 166, 184, 213, 237, 261, 292, 309]
mark_ls_h = [250, 150, 250, 150, 250, 250, 150, 250, 150, 250, 150, 250]

octave = 332
set_time = 1000

finger_num = 1

capture = cv2.VideoCapture(1)

capture.set(3, 640) #가로길이
capture.set(4, 480) #세로길이

img_counter_gen = 0
frame_set = []
start_time = time.time()


final_position = [[-1 for col in range(finger_num*2)] for row in range(10000)]

count = 6001

##### 이미지 출력 #####

def start_show_img():
    print("start_show_img")

    global start_time

    global count
    while count <= 6100: ##################

        left_hand = random.randrange(0, 36)

        print("none")

        final_position[num_imgs][0] = left_hand

        ##### 이미지 출력 #####

        img1 = cv2.imread('test/piano3.png') #piano image

        # 왼손
        # 옥타브
        #print("left")

        k = left_hand

        mark_w = octave * (k // 12) + mark_ls_w[k % 12]
        mark_h = mark_ls_h[k % 12]
        cv2.circle(img1, (mark_w, mark_h), 10, (0, 0, 255), -1) # 빨강

        start_time = time.time()

        cv2.imshow('img', img1)

        video_capture()  #####

        # data_save(num_imgs)

        cv2.destroyWindow('img') ##

        count += 1


def video_capture():

    global start_time, img_counter_gen , num_imgs, final_position
    while True:
        # print("video")
        ret, frame = capture.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        cv2.imshow('frame', gray)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            exit()

        if time.time() - start_time >= set_time:
            img_name = "./sample/piano_1_total/{}_a_{}.jpg".format(count, 36) ###################
            cv2.imwrite(img_name, frame)
            print("{} written!".format(img_counter_gen))
            img_counter_gen += 1
            num_imgs += 1
            break


import os

dir = 'C:/Users/ujin4/PycharmProjects/untitled/sample/piano_1_total'###################

num_imgs = 0
for pack in os.walk(dir):
    for _ in pack[2]:
        num_imgs += 1
    #print (num_imgs)


'''
###### 엑셀에 데이터 정답 저장 #####
from openpyxl import load_workbook

excel_filename = './xlsx/PianoData1_total.xlsx' ###################

# 엑셀 파일 열기
wb = load_workbook(filename = excel_filename)
# 현재 Active Sheet 열기
ws = wb.active


def data_save(num_imgs):
    global final_position
    ws.cell(row=num_imgs-1 + 3, column = 1).value = num_imgs - 1 # 좌측 번호

    # print("final_position[num_imgs]: ", final_position[num_imgs-1] ,end="")
    for i in range(0, finger_num*2):
        if final_position[num_imgs-1][i] != -1:
            ws.cell(row = num_imgs-1+3, column = 38).value = '1' ################### row 잘못됬으면 바꾸기
    wb.save(filename='./xlsx/PianoData1_total.xlsx') ###################
    print("")
'''

start_show_img()