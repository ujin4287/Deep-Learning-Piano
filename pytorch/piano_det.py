import os
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
from torch.utils.data import DataLoader
import torchvision
import torchvision.transforms as transforms
from openpyxl import load_workbook
import re
import numpy

from sklearn.metrics import accuracy_score, recall_score, precision_score, f1_score
from resnet2 import resnet

import PIL
import matplotlib.pyplot as plt
TransPIL = transforms.ToPILImage(mode='')

def img_show(image):

    z = image * torch.tensor([0.5, 0.5, 0.5]).view(3, 1, 1).to(device)
    z = z + torch.tensor([0.5, 0.5, 0.5]).view(3, 1, 1).to(device)

    img2 = transforms.ToPILImage(mode='RGB')(z)
    plt.imshow(img2)


#label_ls_batch = []  # ex) [batch_idx][16][36]

######################################################################

def initialize_train():
    global label_ls_train
    num_keys = 37

    ###  label을 저장한 파일 개수 확인 ###
    dir = 'C:/Users/ujin4/PycharmProjects/untitled/augmentation/sample/total' ################변경

    num_imgs = 0

    for pack in os.walk(dir):
        for _ in pack[2]:
            num_imgs += 1

    ###### 엑셀에서 label 확인 #####
    excel_filename = './xlsx/PianoData1_total.xlsx'
    # 엑셀 파일 열기
    load_wb = load_workbook(filename=excel_filename)
    # 현재 Active Sheet 열기
    load_ws = load_wb['Sheet1']

    label_ls_train = torch.zeros(num_imgs, num_keys)  # ex) [200][36]

    for i in range(0, num_imgs):
        # print("file :", load_ws.cell(i+3, 1).value, end=" ")
        for j in range(0, 37):
            if load_ws.cell(i + 3, j + 2).value == '1':
                # print(type(load_ws.cell(1, j + 2).value))
                label_ls_train[i][j] = 1
    print("train_",num_imgs)


######################################################################

def initialize_test():
    global label_ls_test
    num_keys = 37

    ###  label을 저장한 파일 개수 확인 ###
    dir = 'C:/Users/ujin4/PycharmProjects/untitled/augmentation/sample_test/total' ################변경

    num_imgs = 0

    for pack in os.walk(dir):
        for _ in pack[2]:
            num_imgs += 1

    ###### 엑셀에서 label 확인 #####
    excel_filename = './xlsx/PianoData1_total.xlsx'
    # 엑셀 파일 열기
    load_wb = load_workbook(filename=excel_filename)
    # 현재 Active Sheet 열기
    load_ws = load_wb['Sheet1']

    label_ls_test = torch.zeros(num_imgs, num_keys)  # ex) [200][36]

    for i in range(0, num_imgs):
        # print("file :", load_ws.cell(i+3, 1).value, end=" ")
        for j in range(0, 37):
            if load_ws.cell(i + 3, j + 2).value == '1':
                # print(type(load_ws.cell(1, j + 2).value))
                label_ls_test[i][j] = 1
    print("test_",num_imgs)

######################################################################

from PIL import Image
from torch.utils.data import Dataset


class CustomImageDataset_train(Dataset):
    def read_data_set(self):

        all_img_files = []
        all_labels = []
        initialize_train()
        label = label_ls_train

        class_names = os.walk(self.data_set_path).__next__()[1] # 클래스별 분류되어 있는 파일의 이름을 불러옴

        print("CustomImageDataset_train")

        for index, class_name in enumerate(class_names):

            # label = index # 이부분을 수정해야 할듯?
            # print("label", label)

            img_dir = os.path.join(self.data_set_path, class_name) # os.path.join('C:\Tmp', 'a', 'b') --> "C:\Tmp\a\b" # 즉, 여기서는 클래스 이름별 경로를 탐색
            img_files = os.walk(img_dir).__next__()[2] # 위의 경로를 이용한 파일 불러오기

            for img_file in img_files: # 이미지를 하나씩 불러옴

                numbers = re.findall("\d+", img_file)
                number = int(numbers[0])
                # print("train",number)

                img_file = os.path.join(img_dir, img_file)
                img = Image.open(img_file) # 이미지 파일 읽음

                if img is not None:
                    all_img_files.append(img_file) # 각 파일을 결합
                    all_labels.append(label[number]) # 이부분 수정하면될듯

        return all_img_files, all_labels, len(all_img_files), len(class_names)

    def __init__(self, data_set_path, transforms=None):
        self.data_set_path = data_set_path
        self.image_files_path, self.labels, self.length, self.num_classes = self.read_data_set()
        self.transforms = transforms

    def __getitem__(self, index):
        image = Image.open(self.image_files_path[index])
        image = image.convert("RGB")

        if self.transforms is not None:
            image = self.transforms(image)

        return {'image': image, 'label': self.labels[index]}

    def __len__(self):
        return self.length

class CustomImageDataset_test(Dataset):
    def read_data_set(self):

        all_img_files = []
        all_labels = []
        initialize_test()
        label = label_ls_test

        class_names = os.walk(self.data_set_path).__next__()[1] # 클래스별 분류되어 있는 파일의 이름을 불러옴

        print("CustomImageDataset_test")

        for index, class_name in enumerate(class_names):

            # label = index # 이부분을 수정해야 할듯?
            # print("label", label)

            img_dir = os.path.join(self.data_set_path, class_name) # os.path.join('C:\Tmp', 'a', 'b') --> "C:\Tmp\a\b" # 즉, 여기서는 클래스 이름별 경로를 탐색
            img_files = os.walk(img_dir).__next__()[2] # 위의 경로를 이용한 파일 불러오기

            for img_file in img_files: # 이미지를 하나씩 불러옴

                numbers = re.findall("\d+", img_file)
                number = int(numbers[0])
                # print("test",number)

                img_file = os.path.join(img_dir, img_file)
                img = Image.open(img_file) # 이미지 파일 읽음

                if img is not None:
                    all_img_files.append(img_file) # 각 파일을 결합
                    all_labels.append(label[number]) # 이부분 수정하면될듯

        return all_img_files, all_labels, len(all_img_files), len(class_names)

    def __init__(self, data_set_path, transforms=None):
        self.data_set_path = data_set_path
        self.image_files_path, self.labels, self.length, self.num_classes = self.read_data_set()
        self.transforms = transforms

    def __getitem__(self, index):
        image = Image.open(self.image_files_path[index])
        image = image.convert("RGB")

        if self.transforms is not None:
            image = self.transforms(image)

        return {'image': image, 'label': self.labels[index]}

    def __len__(self):
        return self.length

import torch.nn.functional as F

class FocalLoss(nn.Module):
    def __init__(self, alpha=1, gamma=2, logits=False, reduce=True):
        super(FocalLoss, self).__init__()
        self.alpha = alpha
        self.gamma = gamma
        self.logits = logits
        self.reduce = reduce

    def forward(self, inputs, targets):
        if self.logits:
            BCE_loss = F.binary_cross_entropy_with_logits(inputs, targets, reduce=False)
        else:
            BCE_loss = F.binary_cross_entropy(inputs, targets, reduce=False)
        pt = torch.exp(-BCE_loss)
        F_loss = self.alpha * (1-pt)**self.gamma * BCE_loss

        if self.reduce:
            return torch.mean(F_loss)
        else:
            return F_loss

transforms_train = transforms.Compose([
    transforms.Resize((360, 360)),
    transforms.ToTensor(),
    transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5)),
])
transforms_test = transforms.Compose([
    transforms.Resize((360, 360)),
    transforms.ToTensor(),
    transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5)),
])
train_set = CustomImageDataset_train(data_set_path="C:/Users/ujin4/PycharmProjects/untitled/augmentation/sample/total", transforms=transforms_train)
test_set = CustomImageDataset_test(data_set_path="C:/Users/ujin4/PycharmProjects/untitled/augmentation/sample_test/total", transforms=transforms_test)


# print(train_set.__getitem__(20)) #이미지를 가져온다.
# print(len(train_set)) # 개수 현재는 200

batch_size = 10

train_loader = DataLoader(dataset=train_set,
                         batch_size=batch_size,
                         shuffle=True,
                         num_workers=4)

test_loader = DataLoader(dataset=test_set,
                         batch_size=batch_size,
                         shuffle=False,
                         num_workers=4)


device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

net = resnet()

if torch.cuda.device_count() > 1:
    net = nn.DataParallel(net)
net = net.to(device)

criterion = FocalLoss() ## FocalLoss 사용
optimizer = optim.Adam(net.parameters(), lr=1e-3, weight_decay=1e-5)
step_lr_scheduler = lr_scheduler.MultiStepLR(optimizer, milestones=[20, 30, 40], gamma=0.5)


# visdom 이용
# python -m visdom server
import visdom
import numpy as np

acc_train = torch.zeros(1, 1)
acc_test = torch.zeros(1, 1)


# 학습
def train(epoch, total_step_train):
    net.train()
    global acc_train

    train_loss = 0
    correct = 0
    total = 0

    temp1 = []
    temp2 = []

    total_inputs_num = 0

    for batch_idx, item in enumerate(train_loader):

        inputs = item['image'].to(device)
        targets = item['label'].to(device)

        # print("input:", inputs[0])
        # print("targets:", targets[0])
        # temp = targets[0].tolist()
        # print("index:", temp.index(1))
        # print(inputs.shape) # torch.Size([3, 3, 480, 480])
        # print(targets.shape) # torch.Size([3, 36])

        #inputs_num = inputs.size()[0]

        ### 이미지 확인 ###
        # img_show(inputs[0])

        inputs = inputs.to(device)
        targets = targets.to(device)

        outputs = net(inputs)

        # print(outputs)
        # print(outputs.shape)

        loss = criterion(outputs, targets)

        # print("inputs_num",inputs_num)
        # print("total_inputs_num",total_inputs_num)

        # print("input:", inputs.shape) # image # input: torch.Size([16, 3, 500, 500])
        # print("targets:",batch_idx , targets.shape) # targets: 8 tensor([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], device='cuda:0') # targets: 8 torch.Size([16],[36])
        # if batch_idx == 9:
        #    print(targets)

        optimizer.zero_grad()  # 초기화
        loss.backward()  # 역전파 단계: 모델의 매개변수에 대한 손실의 변화도를 계산합니다
        optimizer.step()  # Optimizer의 step 함수를 호출하면 매개변수가 갱신됩니다.

        step_lr_scheduler.step() ### 위치?

        train_loss += loss.item()
        # print("outputs:", outputs)
        # print("outputs:", outputs.shape) # outputs: torch.Size([3, 36])
        # print("targets:", targets)
        total += targets.size(0) * 37 # none 포함
        # print("total num images:", total / 36)

        output = (outputs > 0.5).float()
        correct += (output == targets).float().sum()

        '''
        print(batch_idx)
        print("outputs:  ", outputs[0])
        print("output:  ", output[0])
        print("targets: ", targets[0])
        '''
        '''
        print("output:  ", output[0])
        print("targets: ", targets[0])
        print("output:  ", output[1])
        print("targets: ", targets[1])
        print("output:  ", output[2])
        print("targets: ", targets[2])

        print("------train f1 score---------")
        print("None", f1_score(targets.tolist(), output.tolist(), average=None))
        print("mirco", f1_score(targets.tolist(), output.tolist(), average='micro'))  # f1_score(y_true, y_pred, average='micro')
        print("macro", f1_score(targets.tolist(), output.tolist(), average='macro'))
        print("weighted", f1_score(targets.tolist(), output.tolist(), average='weighted'))
        print("sample", f1_score(targets.tolist(), output.tolist(), average='samples'))
        print("-----------------------------")

        '''

        acc = 100. * correct / total

        # print(total_step_train)

        temp1.extend(targets.tolist())
        temp2.extend(output.tolist())

        # temp1_tensor = torch.tensor(temp1)
        # temp2_tensor = torch.tensor(temp2)

        # print("temp1_tensor shape", temp1_tensor.shape) # append # temp1_tensor shape torch.Size([1, 3, 36]) # extend # temp1_tensor shape torch.Size([210, 36])


    acc = 100 * correct / total
    acc_train = acc

    print('train epoch : {} [{}/{}]| loss: {:.3f} | acc: {:.3f}'.format(
        epoch, batch_idx, len(train_loader), train_loss / (batch_idx + 1), acc))


    total_step_train += 1
    # print(total_step_train)

    # temp1 = numpy.array(temp1).cpu()
    # temp2 = numpy.array(temp2).cpu()

    # temp1_tensor = torch.tensor(temp1)
    # temp2_tensor = torch.tensor(temp2)

    # print(temp1_tensor.shape) #torch.Size([730, 36])
    # print(temp2_tensor.shape)

    f1_score_None = f1_score(temp1, temp2, average=None)
    f1_score_micro = f1_score(temp1, temp2, average='micro')
    f1_score_macro = f1_score(temp1, temp2, average='macro')
    f1_score_weighted = f1_score(temp1, temp2, average='weighted')
    f1_score_samples = f1_score(temp1, temp2, average='samples')

    print("train f1 score")
    print("None", f1_score_None)
    print("micro", f1_score_micro) # f1_score(y_true, y_pred, average='micro')
    print("macro", f1_score_macro)
    print("weighted", f1_score_weighted)
    print("samples", f1_score_samples)
    print("-------------------------------------")

    vis.line(Y=torch.tensor([acc]),
             X=torch.tensor([epoch]),
             win=plot1,
             name='train_acc',
             update='append')

    vis.line(Y=torch.tensor([train_loss]),
             X=torch.tensor([epoch]),
             win=plot2,
             name='train_loss',
             update='append')

    vis.line(Y=torch.tensor([f1_score_micro]),
             X=torch.tensor([epoch]),
             win=plot3,
             name='train_f1_micro',
             update='append')

    vis.line(Y=torch.tensor([f1_score_weighted]),
             X=torch.tensor([epoch]),
             win=plot4,
             name='train_f1_weighted',
             update='append')

    vis.line(Y=torch.tensor([f1_score_samples]),
             X=torch.tensor([epoch]),
             win=plot5,
             name='train_f1_samples',
             update='append')

    return total_step_train


def test(epoch, total_step_test, best_acc):
    net.eval()

    global acc_test

    test_loss = 0
    correct = 0
    total = 0

    temp1 = []
    temp2 = []

    with torch.no_grad():

        total_inputs_num = 0

        for batch_idx, item in enumerate(test_loader):
            inputs = item['image'].to(device)
            targets = item['label'].to(device)

            # print("input:", inputs[0])
            # print("targets:", targets[0])
            # temp = targets[0].tolist()
            # print("index:", temp.index(1))
            # print(inputs.shape) # torch.Size([3, 3, 500, 500])
            # print(targets.shape) # torch.Size([3, 146, 36])

            # inputs_num = inputs.size()[0]

            ### 이미지 확인 ###
            # img_show(inputs[0])

            inputs = inputs.to(device)
            targets = targets.to(device)

            outputs = net(inputs)

            # print(outputs)
            # print(outputs.shape)

            loss = criterion(outputs, targets)

            # print("inputs_num",inputs_num)
            # print("total_inputs_num",total_inputs_num)

            # print("input:", inputs.shape) # image # input: torch.Size([16, 3, 500, 500])
            # print("targets:",batch_idx , targets.shape) # targets: 8 tensor([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], device='cuda:0') # targets: 8 torch.Size([16],[36])
            # if batch_idx == 9:
            #    print(targets)

            test_loss += loss.item()
            # print("outputs:", outputs)
            # print("targets:", targets)
            total += targets.size(0) * 37
            # print("total num images:", total / 36)

            # 1
            output = (outputs > 0.5).float()
            correct += (output == targets).float().sum()
            # 2
            '''
            for i in range(0, targets.size(0)):
                if output[i] == targets[i] and targets == 1:
                    correct += (output == targets).float().sum()
            '''

            # print("output:", output)
            # print("correct:", correct)

            acc = 100. * correct / total

            temp1.extend(targets.tolist())
            temp2.extend(output.tolist())


        acc = 100 * correct / total
        acc_test = acc
        print('test epoch : {} [{}/{}]| loss: {:.3f} | acc: {:.3f}'.format(
            epoch, batch_idx, len(test_loader), test_loss / (batch_idx + 1), acc))

        '''
        vis.line(Y=torch.tensor([acc]),
                             X=torch.tensor([total_step_test]),
                             win=plot3,
                             name = 'test_step',
                             update='append')  # accuracy를 구하는 수식을 Y값으로 epoch를 X값으로
                    total_step_test += 1
                    '''
        '''
        vis.line(Y=torch.tensor([acc]),
                 X=torch.tensor([epoch]),
                 win=plot,
                 name='test',
                 update='append')  # accuracy를 구하는 수식을 Y값으로 epoch를 X값으로
        '''

    f1_score_None = f1_score(temp1, temp2, average=None)
    f1_score_micro = f1_score(temp1, temp2, average='micro')
    f1_score_macro = f1_score(temp1, temp2, average='macro')
    f1_score_weighted = f1_score(temp1, temp2, average='weighted')
    f1_score_samples = f1_score(temp1, temp2, average='samples')

    print("test f1 score")
    print("None", f1_score_None)
    print("micro", f1_score_micro)  # f1_score(y_true, y_pred, average='micro')
    print("macro", f1_score_macro)
    print("weighted", f1_score_weighted)
    print("samples", f1_score_samples)
    print("-------------------------------------")

    vis.line(Y=torch.tensor([acc_test]),
             X=torch.tensor([epoch]),
             win=plot1,
             name='test_acc',
             update='append')

    vis.line(Y=torch.tensor([test_loss]),
             X=torch.tensor([epoch]),
             win=plot2,
             name='test_loss',
             update='append')

    vis.line(Y=torch.tensor([f1_score_micro]),
             X=torch.tensor([epoch]),
             win=plot3,
             name='test_f1_micro',
             update='append')

    vis.line(Y=torch.tensor([f1_score_weighted]),
             X=torch.tensor([epoch]),
             win=plot4,
             name='test_f1_weighted',
             update='append')

    vis.line(Y=torch.tensor([f1_score_samples]),
             X=torch.tensor([epoch]),
             win=plot5,
             name='test_f1_samples',
             update='append')


    if acc > best_acc:
        print('==> Saving model..')
        state = {
            'net': net.state_dict(),
            'acc': acc,
            'epoch': epoch,
            'optimizer_state_dict': optimizer.state_dict(),
            'loss': test_loss,
            'f1_score_None': f1_score_None,
            'f1_score_micro': f1_score_micro,
            'f1_score_weighted': f1_score_weighted,
            'f1_score_samples': f1_score_samples,
        }
        if not os.path.isdir('save_model'):
            os.mkdir('save_model')
        torch.save(state, './save_model/piano_model.pth')
        best_acc = acc

    return best_acc, total_step_test

if __name__ == '__main__':

    best_acc = 0
    epoch = 0
    total_step_train = 0
    total_step_test = 0

    vis = visdom.Visdom()

    plot1 = vis.line(Y=torch.tensor([0]), X=torch.tensor([0])) # acc
    plot2 = vis.line(Y=torch.tensor([0]), X=torch.tensor([0])) # loss
    plot3 = vis.line(Y=torch.tensor([0]), X=torch.tensor([0])) # f1_score_micro
    plot4 = vis.line(Y=torch.tensor([0]), X=torch.tensor([0])) # f1_score_weighted
    plot5 = vis.line(Y=torch.tensor([0]), X=torch.tensor([0])) # f1_score_samples

    ### 이어서 학습하기??

    '''
    if os.path.isfile('./save_model/piano_model.pth'):
        checkpoint = torch.load('./save_model/piano_model.pth')
        net.load_state_dict(checkpoint['net'])
        optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
        epoch = checkpoint['epoch']
        loss = checkpoint['loss']
        print("===============================")
        print("==========load model===========")
        print("===============================")
        # print(checkpoint)
        # print(net)
        print("epoch:", epoch)
        print("loss", loss)
        print("optimizer",optimizer)
    else:
        print("There is no save_model")
    '''

    while True:
        epoch += 1
        total_step_train = train(epoch, total_step_train)
        best_acc, total_step_test = test(epoch, total_step_test, best_acc)

        print('best test accuracy is ', best_acc)

        if epoch >= 50:
            break




