import torch
import os

dir = 'C:/Users/ujin4/PycharmProjects/untitled/sample'

num_imgs = 0
for pack in os.walk(dir):
    for _ in pack[2]:
        num_imgs += 1
    #print (num_imgs)


###### 엑셀 label 확인 #####
from openpyxl import load_workbook

excel_filename = './xlsx/PianoData2.xlsx'

# 엑셀 파일 열기
load_wb = load_workbook(filename = excel_filename)
# 현재 Active Sheet 열기
load_ws = load_wb['Sheet1']

label_ls = torch.zeros(num_imgs, 36)

for i in range(0, num_imgs):
    # print("file :", load_ws.cell(i+3, 1).value, end=" ")
    for j in range(0, 36):
        if load_ws.cell(i + 3, j + 2).value == '1':
            # print(type(load_ws.cell(1, j + 2).value))
            label_ls[i][j] = 1
            # print(load_ws.cell(1, j + 2).value, end=" ")
    # print("")
    # print(label_ls[i])