import os
from PIL import Image

import torch
from torch.utils.data import Dataset, DataLoader
from torch import nn
from torchvision import transforms

import torch
import os

#####
num_epochs = 10
#####
dir = 'C:/Users/ujin4/PycharmProjects/untitled/sample'

num_imgs = 0
for pack in os.walk(dir):
    for _ in pack[2]:
        num_imgs += 1
    #print (num_imgs)


###### 엑셀 label 확인 #####
from openpyxl import load_workbook

excel_filename = './xlsx/PianoData2.xlsx'

# 엑셀 파일 열기
load_wb = load_workbook(filename = excel_filename)
# 현재 Active Sheet 열기
load_ws = load_wb['Sheet1']

label_ls = torch.zeros(num_imgs, 36)

for i in range(0, num_imgs):
    # print("file :", load_ws.cell(i+3, 1).value, end=" ")
    for j in range(0, 36):
        if load_ws.cell(i + 3, j + 2).value == '1':
            # print(type(load_ws.cell(1, j + 2).value))
            label_ls[i][j] = 1
            # print(load_ws.cell(1, j + 2).value, end=" ")
    # print("")
    # print(label_ls[i])


######

class CustomImageDataset(Dataset):
    def read_data_set(self):

        all_img_files = []
        all_labels = []

        class_names = os.walk(self.data_set_path).__next__()[1] # 클래스별 분류되어 있는 파일의 이름을 불러옴

        for index, class_name in enumerate(class_names):


            label = index # 이부분을 수정해야 할듯?
            print("label",label)

            label = label_ls
            print("label",label)

            img_dir = os.path.join(self.data_set_path, class_name) # os.path.join('C:\Tmp', 'a', 'b') --> "C:\Tmp\a\b" # 즉, 여기서는 클래스 이름별 경로를 탐색
            img_files = os.walk(img_dir).__next__()[2] # 위의 경로를 이용한 파일 불러오기

            for img_file in img_files: # 이미지를 하나씩 불러옴
                img_file = os.path.join(img_dir, img_file)
                img = Image.open(img_file) # 이미지 파일 읽음
                print(img)

                if img is not None:
                    all_img_files.append(img_file) # 각 파일을 결합
                    all_labels.append(label) # 이부분 수정하면될듯

        return all_img_files, all_labels, len(all_img_files), len(class_names)

    def __init__(self, data_set_path, transforms=None):
        self.data_set_path = data_set_path
        self.image_files_path, self.labels, self.length, self.num_classes = self.read_data_set()
        self.transforms = transforms

    def __getitem__(self, index):
        image = Image.open(self.image_files_path[index])
        image = image.convert("RGB")

        if self.transforms is not None:
            image = self.transforms(image)

        return {'image': image, 'label': self.labels[index]}

    def __len__(self):
        return self.length


transforms_train = transforms.Compose([
    transforms.Resize((500, 500)),
    transforms.ToTensor(),
    transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5)),
])

batch_size = 3

train_data_set = CustomImageDataset(data_set_path="./sample/piano_1", transforms=transforms_train)

'''
train_loader = DataLoader(dataset = train_data_set,
                         batch_size = batch_size,
                         shuffle = True,
                         num_workers = 4)
'''





