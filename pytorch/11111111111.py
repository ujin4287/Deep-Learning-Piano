
from openpyxl import load_workbook

excel_filename = './PianoData.xlsx'
wb = load_workbook(filename = excel_filename)
ws = wb.active

finalposition=[1,4,10]
num_imgs=1

def data_save(finalposition, num_imgs):
    ws.cell(row=num_imgs + 2, column = 1).value = num_imgs - 1
    for i in finalposition:
        ws.cell(row = num_imgs+2, column = i+2).value = '1' ################### row 잘못됬으면 바꾸기
    wb.save(filename='./xlsx/PianoData1.xlsx')

data_save(finalposition,num_imgs)